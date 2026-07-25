"""
gnina_vs.reporter
======================
Step 4 of the pipeline: parse Gnina scores (from stdout or, as a fallback,
from the docked SDF's embedded properties), export a ranked CSV/Excel
report, merge all docked poses into a single ranked SDF, and support
recovery (re-scoring from existing docked SDFs without re-docking).
"""

from __future__ import annotations

import logging
import os
import re
from pathlib import Path

import pandas as pd
from tqdm import tqdm

from rdkit import Chem

logger = logging.getLogger(__name__)

# Header keyword patterns
_HDR_AFFINITY = re.compile(r"affinity", re.IGNORECASE)
_HDR_CNN_SCORE = re.compile(r"cnn.?score", re.IGNORECASE)
_HDR_CNN_AFFINITY = re.compile(r"cnn.?affinity", re.IGNORECASE)

_KV_AFFINITY = re.compile(r"affinity\s*[=:]\s*([-\d.]+)", re.IGNORECASE)
_KV_CNN_SCORE = re.compile(r"cnn.?score\s*[=:]\s*([\d.]+)", re.IGNORECASE)
_KV_CNN_AFFINITY = re.compile(r"cnn.?affinity\s*[=:]\s*([\d.]+)", re.IGNORECASE)


def _split_pipe_row(line: str) -> list[str]:
    parts = [p.strip() for p in line.split("|")]
    return [p for p in parts if p]


def _parse_gnina_log(raw_log: str) -> dict[str, float | None]:
    """Parse Gnina stdout to extract mode-1 scores.

    Gnina v1.3.2 produces this format::

        mode |  affinity  |  intramol  |    CNN     |   CNN
             | (kcal/mol) | (kcal/mol) | pose score | affinity
        -----+------------+------------+------------+----------
            1       -8.32       -0.34       0.7574      6.438

    Key quirks handled here:
    - The column headers span TWO lines (affinity on line 1, CNN scores on
      line 2).
    - Data rows are space-separated, NOT pipe-separated.
    - The separator line starts with "---".
    """
    scores: dict[str, float | None] = {
        "vina_affinity": None,
        "cnn_score": None,
        "cnn_affinity": None,
    }

    if not raw_log or raw_log == "TIMEOUT":
        return scores

    lines = raw_log.splitlines()

    # -- Strategy 1: find the "-----+" separator, then parse the next line --
    # This is the most reliable approach for gnina's space-delimited table.
    for i, line in enumerate(lines):
        if line.strip().startswith("---") and "+" in line:
            for data_line in lines[i + 1:]:
                stripped = data_line.strip()
                if not stripped:
                    continue
                parts = stripped.split()
                # Expect: mode affinity intramol cnn_pose_score cnn_affinity
                if len(parts) >= 5 and parts[0] == "1":
                    try:
                        scores["vina_affinity"] = float(parts[1])
                        scores["cnn_score"] = float(parts[3])
                        scores["cnn_affinity"] = float(parts[4])
                        return scores
                    except (IndexError, ValueError):
                        pass
                break  # only want mode 1; stop after first data row
            break

    # -- Strategy 2: key=value fallback (older gnina versions) --
    m_aff = _KV_AFFINITY.search(raw_log)
    m_cs = _KV_CNN_SCORE.search(raw_log)
    m_ca = _KV_CNN_AFFINITY.search(raw_log)
    if m_aff or m_cs or m_ca:
        if m_aff:
            scores["vina_affinity"] = float(m_aff.group(1))
        if m_cs:
            scores["cnn_score"] = float(m_cs.group(1))
        if m_ca:
            scores["cnn_affinity"] = float(m_ca.group(1))
        return scores

    # -- Strategy 3: pipe-delimited table (some gnina builds) --
    for i, line in enumerate(lines):
        if _HDR_AFFINITY.search(line) and _HDR_CNN_SCORE.search(line):
            for data_line in lines[i + 1:]:
                if not data_line.strip() or data_line.strip().startswith("-"):
                    continue
                parts = _split_pipe_row(data_line)
                if not parts or parts[0] != "1":
                    continue
                try:
                    scores["vina_affinity"] = float(parts[1])
                    scores["cnn_score"] = float(parts[-2])
                    scores["cnn_affinity"] = float(parts[-1])
                    return scores
                except (IndexError, ValueError):
                    break
            break

    return scores


def _score_from_sdf(sdf_path: str) -> dict[str, float | None]:
    """Fallback: read scores from the docked SDF's embedded SD properties
    (``minimizedAffinity``, ``CNNscore``, ``CNNaffinity``), used when the
    stdout log didn't contain them (e.g. captured before writing completed)."""
    scores: dict[str, float | None] = {
        "vina_affinity": None, "cnn_score": None, "cnn_affinity": None,
    }
    if not sdf_path or not os.path.exists(sdf_path):
        return scores
    try:
        suppl = Chem.SDMolSupplier(sdf_path, removeHs=False)
        first = next((m for m in suppl if m is not None), None)
        if first is None:
            return scores

        def _prop(mol, *names):
            for n in names:
                if mol.HasProp(n):
                    try:
                        return float(mol.GetProp(n))
                    except ValueError:
                        pass
            return None

        scores["vina_affinity"] = _prop(
            first, "minimizedAffinity", "affinity", "docking_score", "Affinity")
        scores["cnn_score"] = _prop(first, "CNNscore", "cnn_score", "CNN_score")
        scores["cnn_affinity"] = _prop(first, "CNNaffinity", "cnn_affinity", "CNN_affinity")
    except Exception:
        pass
    return scores


def _resolve_top_n(top_n: int | str, total: int) -> int:
    """Interpret the --top-n CLI value: 0 or 'all' means export everything."""
    if isinstance(top_n, str) and top_n.strip().lower() == "all":
        return total
    top_n = int(top_n)
    return total if top_n <= 0 else top_n


def merge_docked_sdf(df: pd.DataFrame, output_sdf: str) -> str:
    """Merge the best (first) pose of each docked ligand into a single,
    rank-ordered multi-molecule SDF for easy visual inspection.

    ``df`` must already be sorted in the desired (best-first) order and
    contain an ``output_sdf`` column pointing at each ligand's docked SDF,
    plus ``ID`` and rank-relevant score columns.
    """
    writer = Chem.SDWriter(output_sdf)
    n_written = 0
    for rank, row in enumerate(df.itertuples(index=False), start=1):
        sdf_path = getattr(row, "output_sdf", None)
        if not sdf_path or not os.path.exists(sdf_path):
            continue
        try:
            suppl = Chem.SDMolSupplier(sdf_path, removeHs=False)
            first = next((m for m in suppl if m is not None), None)
        except Exception:
            first = None
        if first is None:
            continue
        first.SetProp("_Name", f"rank{rank}_{getattr(row, 'ID', rank)}")
        first.SetProp("Rank", str(rank))
        if getattr(row, "vina_affinity", None) is not None:
            first.SetProp("vina_affinity", str(row.vina_affinity))
        if getattr(row, "cnn_score", None) is not None:
            first.SetProp("cnn_score", str(row.cnn_score))
        if getattr(row, "cnn_affinity", None) is not None:
            first.SetProp("cnn_affinity", str(row.cnn_affinity))
        writer.write(first)
        n_written += 1
    writer.close()
    logger.info("Merged %d ranked poses into '%s'.", n_written, output_sdf)
    return output_sdf


def extract_results(
    docking_results: list[dict],
    input_path: str | None = None,
    output_csv: str = "results_summary.csv",
    output_excel: str = "top_results.xlsx",
    top_n: int | str = 100,
    merged_sdf: str | None = None,
    id_col: str = "ID",
) -> pd.DataFrame:
    """Parse Gnina logs, build a results DataFrame, save as CSV, and write
    the top-``top_n`` compounds (by CNN Affinity, then Vina Affinity) to an
    Excel sheet that includes the original ID column from ``input_path``.

    Parameters
    ----------
    docking_results : list[dict]
        Output of `gnina_vs.docking.run_gnina_docking`.
    input_path : str | None
        Path to the original ligand Excel/CSV file (for the ID column).
    output_csv : str
        Destination CSV for the full results table.
    output_excel : str
        Destination Excel file for the exported results (with ID column).
    top_n : int | str
        Number of top-ranked compounds to export. ``0`` or ``"all"``
        exports every processed compound.
    merged_sdf : str | None
        If given, write a single ranked multi-molecule SDF of the exported
        compounds' best poses here.
    id_col : str
        Name of the identifier column in ``input_path``.

    Returns
    -------
    pd.DataFrame
        Full results table (all compounds, sorted).
    """
    logger.info("=" * 60)
    logger.info("STEP 4 -- Results Extraction & Export")
    logger.info("=" * 60)

    # -- Load original input table to recover the ID column --
    id_map: dict[str, str] = {}  # mol_index -> ID
    if input_path and os.path.exists(input_path):
        ext = os.path.splitext(input_path)[1].lower()
        orig_df = pd.read_excel(input_path) if ext in (".xlsx", ".xls") else pd.read_csv(input_path)
        if id_col in orig_df.columns:
            for row_idx, row in orig_df.iterrows():
                id_map[str(row_idx)] = str(row[id_col])
        else:
            logger.warning(
                "  '%s' column not found in '%s'. The ID column in the "
                "output will be populated from the embedded LigandID.",
                id_col, input_path,
            )
    elif input_path:
        logger.warning(
            "  Original input file '%s' not found -- ID column will be "
            "populated from embedded LigandID values.", input_path,
        )

    # -- Build results DataFrame --
    rows = []
    for entry in docking_results:
        parsed = _parse_gnina_log(entry["raw_log"])
        if None in parsed.values():
            fallback = _score_from_sdf(entry.get("output_path", ""))
            for key, val in fallback.items():
                if parsed.get(key) is None:
                    parsed[key] = val

        mol_index = str(entry["index"])
        lig_id = id_map.get(mol_index, entry.get("ligand_id", mol_index))

        rows.append({
            "ID": lig_id,
            "compound_index": entry["index"],
            "mol_name": entry["mol_name"],
            "smiles": entry["smiles"],
            "vina_affinity": parsed["vina_affinity"],
            "cnn_score": parsed["cnn_score"],
            "cnn_affinity": parsed["cnn_affinity"],
            "output_sdf": entry["output_path"],
        })

    df = pd.DataFrame(rows)

    # -- Sort: CNN Affinity descending, then Vina Affinity ascending --
    df.sort_values(
        by=["cnn_affinity", "vina_affinity"],
        ascending=[False, True],
        inplace=True,
        na_position="last",
    )

    # -- Save full results as CSV --
    df.to_csv(output_csv, index=False)
    logger.info("Full results saved to '%s' (%d entries).", output_csv, len(df))

    # -- Export top-N (or all) to Excel --
    successful = df.dropna(subset=["cnn_affinity"])
    n_export = _resolve_top_n(top_n, len(successful))
    export_df = successful.head(n_export).copy()

    # Ensure 'ID' is the first column
    cols = ["ID"] + [c for c in export_df.columns if c != "ID"]
    export_df = export_df[cols]

    sheet_name = "AllResults" if n_export == len(successful) else f"Top{n_export}"
    export_df.to_excel(output_excel, index=False, sheet_name=sheet_name)

    # Auto-fit column widths for readability
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.load_workbook(output_excel)
        ws = wb.active
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)
        wb.save(output_excel)
    except Exception as exc:
        logger.debug("  Column auto-fit skipped: %s", exc)

    logger.info(
        "%d compounds saved to '%s' (sheet: '%s').",
        len(export_df), output_excel, sheet_name,
    )

    if merged_sdf:
        merge_docked_sdf(export_df, merged_sdf)

    # -- Console summary --
    logger.info("Successfully scored: %d/%d compounds.", len(successful), len(df))
    if not successful.empty:
        best = successful.iloc[0]
        logger.info(
            "  Top compound  ID=%s  |  CNN Affinity: %.3f  |  CNN Score: %.3f  |  "
            "Vina: %.3f kcal/mol",
            best["ID"], best["cnn_affinity"], best["cnn_score"], best["vina_affinity"],
        )

    return df


# --------------------------------------------------------------------------- #
# Recovery -- re-score from existing docked SDF files
# --------------------------------------------------------------------------- #
def recover_results(
    ligands_sdf: str,
    out_dir: Path | str = "docking_outputs",
    input_path: str | None = None,
    output_csv: str = "results_summary.csv",
    output_excel: str = "top_results.xlsx",
    top_n: int | str = 100,
    merged_sdf: str | None = None,
    id_col: str = "ID",
) -> pd.DataFrame:
    """Re-extract results from already-docked SDF files without re-running
    Gnina.

    Use this when docking completed successfully but score extraction
    needs to be re-run (e.g. after fixing a log-parsing issue), or simply
    to regenerate a report with different `--top-n` / export settings.
    Reads the compound list from the prepared ligands SDF, locates each
    ``<mol_name>_docked.sdf`` in ``out_dir``, and reads scores from the SDF
    properties Gnina embeds in every output file.
    """
    logger.info("=" * 60)
    logger.info("RECOVERY -- Re-scoring from existing docked SDF files")
    logger.info("=" * 60)

    out_dir = Path(out_dir)

    if not os.path.exists(ligands_sdf):
        raise FileNotFoundError(f"Prepared ligands SDF not found: {ligands_sdf}")

    supplier = Chem.SDMolSupplier(ligands_sdf, removeHs=False)
    molecules = [mol for mol in supplier if mol is not None]
    logger.info("  %d ligands found in '%s'.", len(molecules), ligands_sdf)
    if not molecules:
        raise ValueError(f"No valid molecules found in '{ligands_sdf}'.")

    docking_results = []
    missing = 0
    for mol in tqdm(molecules, desc="Reading docked SDFs", unit="mol", colour="cyan"):
        mol_name = mol.GetProp("_Name") if mol.HasProp("_Name") else "unknown"
        smiles = mol.GetProp("SMILES") if mol.HasProp("SMILES") else ""
        mol_index = mol.GetProp("MolIndex") if mol.HasProp("MolIndex") else "-1"
        lig_id = mol.GetProp("LigandID") if mol.HasProp("LigandID") else mol_index

        out_sdf = out_dir / f"{mol_name}_docked.sdf"
        if not out_sdf.exists():
            missing += 1
            raw_log = "TIMEOUT"
        else:
            raw_log = ""
            try:
                suppl2 = Chem.SDMolSupplier(str(out_sdf), removeHs=False)
                first = next((m for m in suppl2 if m is not None), None)
                if first is not None:
                    aff = first.GetProp("minimizedAffinity") if first.HasProp("minimizedAffinity") else ""
                    cs = first.GetProp("CNNscore") if first.HasProp("CNNscore") else ""
                    ca = first.GetProp("CNNaffinity") if first.HasProp("CNNaffinity") else ""
                    # Reconstruct a minimal gnina-style table so _parse_gnina_log works
                    raw_log = (
                        "mode |  affinity  |  intramol  |    CNN     |   CNN\n"
                        "     | (kcal/mol) | (kcal/mol) | pose score | affinity\n"
                        "-----+------------+------------+------------+----------\n"
                        f"    1    {aff}       0.00    {cs}    {ca}\n"
                    )
            except Exception:
                pass

        docking_results.append({
            "mol_name": mol_name,
            "smiles": smiles,
            "index": mol_index,
            "ligand_id": lig_id,
            "output_path": str(out_dir / f"{mol_name}_docked.sdf"),
            "raw_log": raw_log,
        })

    if missing:
        logger.warning("  %d docked SDF files were missing (will appear as unscored).", missing)

    return extract_results(
        docking_results=docking_results,
        input_path=input_path,
        output_csv=output_csv,
        output_excel=output_excel,
        top_n=top_n,
        merged_sdf=merged_sdf,
        id_col=id_col,
    )
